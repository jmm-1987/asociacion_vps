from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file
from flask_login import login_required, current_user
from models import User, Actividad, Inscripcion, SolicitudSocio, BeneficiarioSolicitud, Beneficiario, db
from datetime import datetime, timedelta
from functools import wraps
import secrets
import string
import re
import unicodedata
import json
import os
import shutil
import subprocess
from io import BytesIO, StringIO
from flask import current_app
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def quitar_acentos(texto):
    """Convierte texto a mayúsculas y quita acentos, pero preserva la ñ"""
    MARKER = '\uE000'  # Carácter privado Unicode que no se usa
    texto = texto.replace('ñ', MARKER).replace('Ñ', MARKER)
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    texto = texto.replace(MARKER, 'Ñ')
    return texto.upper()

admin_bp = Blueprint('admin', __name__)

def directiva_required(f):
    """Decorador para requerir rol de directiva"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_directiva():
            flash('No tienes permisos para acceder a esta página.', 'error')
            return redirect(url_for('socios.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@admin_bp.route('/dashboard')
@login_required
@directiva_required
def dashboard():
    # Socios próximos a vencer (30 días)
    limite_vencimiento = datetime.utcnow() + timedelta(days=30)
    socios_por_vencer = User.query.filter(
        User.rol == 'socio',
        User.fecha_validez <= limite_vencimiento,
        User.fecha_validez > datetime.utcnow()
    ).order_by(User.fecha_validez).all()
    
    # Todas las actividades con número de inscritos
    actividades = Actividad.query.order_by(Actividad.fecha.desc()).all()
    
    # Estadísticas
    total_socios = User.query.filter_by(rol='socio').count()
    total_actividades = Actividad.query.count()
    solicitudes_pendientes = SolicitudSocio.query.filter_by(estado='por_confirmar').count()
    
    return render_template('admin/dashboard.html',
                         socios_por_vencer=socios_por_vencer,
                         actividades=actividades,
                         total_socios=total_socios,
                         total_actividades=total_actividades,
                         solicitudes_pendientes=solicitudes_pendientes)

@admin_bp.route('/socios')
@login_required
@directiva_required
def gestion_socios():
    # Obtener parámetros de búsqueda y filtro
    search_query = request.args.get('search', '').strip()
    solo_ninos = request.args.get('solo_ninos', '').strip() == 'on'
    
    # Construir query base
    query = User.query.filter(User.rol == 'socio')
    
    # Aplicar filtro de solo niños (menores de 18 años)
    if solo_ninos:
        año_actual = datetime.now().year
        año_limite = año_actual - 18
        # Obtener IDs de socios que tienen beneficiarios niños
        socios_ids_con_ninos = db.session.query(Beneficiario.socio_id).filter(
            Beneficiario.ano_nacimiento >= año_limite
        ).distinct().all()
        socios_ids_con_ninos = [s[0] for s in socios_ids_con_ninos]
        
        # Filtrar: socio es niño O tiene beneficiarios niños
        condiciones = [User.ano_nacimiento >= año_limite]
        if socios_ids_con_ninos:
            condiciones.append(User.id.in_(socios_ids_con_ninos))
        
        query = query.filter(db.or_(*condiciones))
    
    # Aplicar búsqueda
    if search_query:
        query = query.filter(
            db.or_(
                User.nombre.contains(search_query),
                User.nombre_usuario.contains(search_query),
                db.func.strftime('%d/%m/%Y', User.fecha_validez).contains(search_query)
            )
        )
    
    socios = query.order_by(User.nombre).all()
    
    # Cargar beneficiarios para cada socio
    for socio in socios:
        socio.beneficiarios_lista = Beneficiario.query.filter_by(socio_id=socio.id).order_by(Beneficiario.nombre).all()
    
    from datetime import datetime as dt
    return render_template('admin/socios.html', socios=socios, search_query=search_query, solo_ninos=solo_ninos, datetime=dt)

@admin_bp.route('/beneficiarios')
@login_required
@directiva_required
def gestion_beneficiarios():
    # Obtener parámetros de búsqueda y filtro
    search_query = request.args.get('search', '').strip()
    solo_ninos = request.args.get('solo_ninos', '').strip() == 'on'
    
    año_actual = datetime.now().year
    año_limite = año_actual - 18
    
    # Obtener beneficiarios tradicionales
    query_beneficiarios = Beneficiario.query.join(User).filter(User.rol == 'socio')
    
    # Aplicar filtro de solo niños (menores de 18 años)
    if solo_ninos:
        query_beneficiarios = query_beneficiarios.filter(Beneficiario.ano_nacimiento >= año_limite)
    
    # Aplicar búsqueda en beneficiarios
    if search_query:
        query_beneficiarios = query_beneficiarios.filter(
            db.or_(
                Beneficiario.nombre.contains(search_query),
                Beneficiario.primer_apellido.contains(search_query),
                Beneficiario.segundo_apellido.contains(search_query),
                Beneficiario.numero_beneficiario.contains(search_query),
                User.nombre.contains(search_query),
                User.numero_socio.contains(search_query)
            )
        )
    
    beneficiarios_list = query_beneficiarios.order_by(Beneficiario.nombre).all()
    
    # Obtener TODOS los socios (todos los socios también son beneficiarios)
    query_socios = User.query.filter(User.rol == 'socio')
    
    # Aplicar filtro de solo niños si está activo
    if solo_ninos:
        query_socios = query_socios.filter(User.ano_nacimiento >= año_limite)
    
    # Aplicar búsqueda en socios
    if search_query:
        query_socios = query_socios.filter(
            db.or_(
                User.nombre.contains(search_query),
                User.nombre_usuario.contains(search_query),
                User.numero_socio.contains(search_query)
            )
        )
    
    socios = query_socios.order_by(User.nombre).all()
    
    # Crear una lista unificada de beneficiarios
    # Convertir beneficiarios a objetos con estructura común
    beneficiarios_unificados = []
    
    # Añadir beneficiarios tradicionales
    for ben in beneficiarios_list:
        ben.socio_info = User.query.get(ben.socio_id)
        ben.es_socio = False
        beneficiarios_unificados.append(ben)
    
    # Añadir TODOS los socios como beneficiarios (cada socio tiene su propia línea)
    # Cada socio aparece como beneficiario de sí mismo, independientemente de si tiene otros beneficiarios
    for socio in socios:
        # Crear un objeto similar a Beneficiario pero para el socio
        # El socio es beneficiario de sí mismo
        class BeneficiarioSocio:
            def __init__(self, socio):
                self.id = socio.id
                # Separar nombre completo en partes
                partes = socio.nombre.split(' ', 2)
                self.nombre = partes[0] if len(partes) > 0 else ''
                self.primer_apellido = partes[1] if len(partes) > 1 else ''
                self.segundo_apellido = partes[2] if len(partes) > 2 else None
                self.ano_nacimiento = socio.ano_nacimiento
                self.numero_beneficiario = socio.numero_socio  # El número de beneficiario es el mismo que el número de socio
                self.fecha_validez = socio.fecha_validez
                self.socio_id = socio.id  # El socio es beneficiario de sí mismo
                self.socio_info = socio  # El socio es su propio socio
                self.es_socio = True
        
        beneficiario_socio = BeneficiarioSocio(socio)
        beneficiarios_unificados.append(beneficiario_socio)
    
    # Ordenar por nombre
    beneficiarios_unificados.sort(key=lambda x: x.nombre)
    
    from datetime import datetime as dt
    return render_template('admin/beneficiarios.html', 
                         beneficiarios=beneficiarios_unificados, 
                         search_query=search_query, 
                         solo_ninos=solo_ninos,
                         datetime=dt,
                         timedelta=timedelta)

@admin_bp.route('/socios/nuevo', methods=['GET', 'POST'])
@login_required
@directiva_required
def nuevo_socio():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        primer_apellido = request.form.get('primer_apellido', '').strip()
        segundo_apellido = request.form.get('segundo_apellido', '').strip()
        movil = request.form.get('movil', '').strip()
        miembros_unidad_familiar = request.form.get('miembros_unidad_familiar', '').strip()
        forma_de_pago = request.form.get('forma_de_pago', '').strip()
        password = request.form.get('password', '').strip()
        ano_nacimiento = request.form.get('ano_nacimiento', '').strip()
        nombre_usuario = request.form.get('nombre_usuario', '').strip()
        
        # Dirección
        calle = request.form.get('calle', '').strip()
        numero = request.form.get('numero', '').strip()
        piso = request.form.get('piso', '').strip()
        poblacion = request.form.get('poblacion', '').strip()
        
        # Validaciones
        if not all([nombre, primer_apellido, movil, miembros_unidad_familiar, forma_de_pago, password, ano_nacimiento, nombre_usuario, calle, numero, poblacion]):
            flash('Todos los campos obligatorios deben estar completos.', 'error')
            from datetime import datetime as dt
            return render_template('admin/nuevo_socio.html', datetime=dt)
        
        # Verificar si el nombre de usuario ya existe
        if User.query.filter_by(nombre_usuario=nombre_usuario).first():
            flash('Ya existe un usuario con este nombre de usuario.', 'error')
            from datetime import datetime as dt
            return render_template('admin/nuevo_socio.html', datetime=dt)
        
        # Validar móvil
        if not re.match(r'^\d{9}$', movil):
            flash('El número de móvil debe tener 9 dígitos.', 'error')
            from datetime import datetime as dt
            return render_template('admin/nuevo_socio.html', datetime=dt)
        
        # Validar año de nacimiento
        try:
            ano_nac = int(ano_nacimiento)
            año_actual = datetime.now().year
            if ano_nac < 1900 or ano_nac > año_actual:
                flash('El año de nacimiento debe estar entre 1900 y el año actual.', 'error')
                from datetime import datetime as dt
                return render_template('admin/nuevo_socio.html', datetime=dt)
            fecha_nacimiento_obj = datetime(ano_nac, 1, 1).date()
        except ValueError:
            flash('Año de nacimiento inválido.', 'error')
            from datetime import datetime as dt
            return render_template('admin/nuevo_socio.html', datetime=dt)
        
        # Validar miembros
        try:
            miembros = int(miembros_unidad_familiar)
            if miembros <= 0:
                raise ValueError()
        except ValueError:
            flash('El número de miembros de la unidad familiar debe ser un número positivo.', 'error')
            from datetime import datetime as dt
            return render_template('admin/nuevo_socio.html', datetime=dt)
        
        # Validar forma de pago
        if forma_de_pago not in ['bizum', 'transferencia']:
            flash('Forma de pago inválida.', 'error')
            from datetime import datetime as dt
            return render_template('admin/nuevo_socio.html', datetime=dt)
        
        # Convertir a mayúsculas y quitar acentos
        nombre = quitar_acentos(nombre)
        primer_apellido = quitar_acentos(primer_apellido)
        if segundo_apellido:
            segundo_apellido = quitar_acentos(segundo_apellido)
        calle = quitar_acentos(calle.upper())
        poblacion = quitar_acentos(poblacion.upper())
        numero = numero.strip()
        piso = piso.strip() if piso else None
        
        # Generar nombre completo
        nombre_completo = f"{nombre} {primer_apellido}"
        if segundo_apellido:
            nombre_completo += f" {segundo_apellido}"
        
        # Fecha de validez siempre al 31/12 del año en curso
        año_actual = datetime.now().year
        fecha_validez = datetime(año_actual, 12, 31, 23, 59, 59)
        
        # Crear nuevo socio
        nuevo_socio = User(
            nombre=nombre_completo,
            nombre_usuario=nombre_usuario,
            rol='socio',
            fecha_alta=datetime.utcnow(),
            fecha_validez=fecha_validez,
            ano_nacimiento=ano_nac,
            fecha_nacimiento=fecha_nacimiento_obj,
            calle=calle,
            numero=numero,
            piso=piso,
            poblacion=poblacion
        )
        nuevo_socio.set_password(password)
        
        try:
            db.session.add(nuevo_socio)
            db.session.commit()
            flash(f'Socio {nombre_completo} registrado exitosamente con validez hasta el 31/12/{año_actual}.', 'success')
            return redirect(url_for('admin.gestion_socios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar el socio: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
            from datetime import datetime as dt
            return render_template('admin/nuevo_socio.html', datetime=dt)
    
    from datetime import datetime as dt
    return render_template('admin/nuevo_socio.html', datetime=dt)

@admin_bp.route('/socios/<int:socio_id>/editar', methods=['GET', 'POST'])
@login_required
@directiva_required
def editar_socio(socio_id):
    socio = User.query.get_or_404(socio_id)
    beneficiarios = Beneficiario.query.filter_by(socio_id=socio.id).order_by(Beneficiario.id).all()
    
    if request.method == 'POST':
        # Obtener datos del formulario
        nombre = request.form.get('nombre', '').strip()
        primer_apellido = request.form.get('primer_apellido', '').strip()
        segundo_apellido = request.form.get('segundo_apellido', '').strip()
        nombre_usuario = request.form.get('nombre_usuario', '').strip()
        ano_nacimiento = request.form.get('ano_nacimiento', '').strip()
        password = request.form.get('password', '').strip()
        numero_socio = request.form.get('numero_socio', '').strip()
        rol = request.form.get('rol', '').strip()
        fecha_alta_str = request.form.get('fecha_alta', '').strip()
        fecha_validez_str = request.form.get('fecha_validez', '').strip()
        fecha_nacimiento_str = request.form.get('fecha_nacimiento', '').strip()
        
        # Dirección
        calle = request.form.get('calle', '').strip()
        numero = request.form.get('numero', '').strip()
        piso = request.form.get('piso', '').strip()
        poblacion = request.form.get('poblacion', '').strip()
        
        # Validaciones básicas
        if not all([nombre, primer_apellido, segundo_apellido, nombre_usuario, calle, numero, poblacion, rol, fecha_validez_str]):
            flash('Todos los campos obligatorios deben estar completos.', 'error')
            from datetime import datetime as dt
            # Separar nombre completo en partes
            partes_nombre = socio.nombre.split(' ', 2)
            nombre_parts = {
                'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
            }
            return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
        
        # Verificar si el nombre de usuario ya existe (excepto para el mismo socio)
        usuario_existente = User.query.filter_by(nombre_usuario=nombre_usuario).first()
        if usuario_existente and usuario_existente.id != socio.id:
            flash('Ya existe un usuario con este nombre de usuario.', 'error')
            from datetime import datetime as dt
            partes_nombre = socio.nombre.split(' ', 2)
            nombre_parts = {
                'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
            }
            return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
        
        # Validar año de nacimiento
        if ano_nacimiento:
            try:
                ano_nac = int(ano_nacimiento)
                año_actual = datetime.now().year
                if ano_nac < 1900 or ano_nac > año_actual:
                    flash('El año de nacimiento debe estar entre 1900 y el año actual.', 'error')
                    from datetime import datetime as dt
                    partes_nombre = socio.nombre.split(' ', 2)
                    nombre_parts = {
                        'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                        'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                        'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
                    }
                    return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
            except ValueError:
                flash('Año de nacimiento inválido.', 'error')
                from datetime import datetime as dt
                partes_nombre = socio.nombre.split(' ', 2)
                nombre_parts = {
                    'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                    'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                    'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
                }
                return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
        else:
            ano_nac = None
        
        # Validar y procesar fecha de nacimiento completa
        fecha_nacimiento_obj = None
        if fecha_nacimiento_str:
            try:
                fecha_nacimiento_obj = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
                # Si hay fecha de nacimiento completa, actualizar también el año
                if not ano_nac:
                    ano_nac = fecha_nacimiento_obj.year
            except ValueError:
                flash('Fecha de nacimiento inválida.', 'error')
                from datetime import datetime as dt
                partes_nombre = socio.nombre.split(' ', 2)
                nombre_parts = {
                    'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                    'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                    'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
                }
                return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
        elif ano_nac:
            # Si solo hay año, crear fecha con 1 de enero
            fecha_nacimiento_obj = datetime(ano_nac, 1, 1).date()
        
        # Validar y procesar fecha de alta
        fecha_alta_obj = None
        if fecha_alta_str:
            try:
                fecha_alta_obj = datetime.strptime(fecha_alta_str, '%Y-%m-%d')
            except ValueError:
                flash('Fecha de alta inválida.', 'error')
                from datetime import datetime as dt
                partes_nombre = socio.nombre.split(' ', 2)
                nombre_parts = {
                    'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                    'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                    'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
                }
                return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
        
        # Validar y procesar fecha de validez
        try:
            fecha_validez_obj = datetime.strptime(fecha_validez_str, '%Y-%m-%dT%H:%M')
        except ValueError:
            flash('Fecha de validez inválida.', 'error')
            from datetime import datetime as dt
            partes_nombre = socio.nombre.split(' ', 2)
            nombre_parts = {
                'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
            }
            return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
        
        # Validar número de socio (debe ser único si se proporciona)
        if numero_socio:
            numero_socio = numero_socio.strip()
            # Verificar que sea único (excepto para el mismo socio)
            socio_existente = User.query.filter_by(numero_socio=numero_socio).first()
            if socio_existente and socio_existente.id != socio.id:
                flash('Ya existe otro socio con este número de socio.', 'error')
                from datetime import datetime as dt
                partes_nombre = socio.nombre.split(' ', 2)
                nombre_parts = {
                    'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                    'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                    'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
                }
                return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
        else:
            numero_socio = None
        
        # Validar rol
        if rol not in ['socio', 'directiva']:
            flash('Rol inválido.', 'error')
            from datetime import datetime as dt
            partes_nombre = socio.nombre.split(' ', 2)
            nombre_parts = {
                'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
            }
            return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
        
        # Convertir a mayúsculas y quitar acentos
        nombre = quitar_acentos(nombre)
        primer_apellido = quitar_acentos(primer_apellido)
        if segundo_apellido:
            segundo_apellido = quitar_acentos(segundo_apellido)
        calle = quitar_acentos(calle.upper())
        poblacion = quitar_acentos(poblacion.upper())
        numero = numero.strip()
        piso = piso.strip() if piso else None
        
        # Generar nombre completo
        nombre_completo = f"{nombre} {primer_apellido}"
        if segundo_apellido:
            nombre_completo += f" {segundo_apellido}"
        
        # Actualizar datos del socio
        socio.nombre = nombre_completo
        socio.nombre_usuario = nombre_usuario
        socio.rol = rol
        socio.ano_nacimiento = ano_nac
        socio.fecha_nacimiento = fecha_nacimiento_obj
        socio.numero_socio = numero_socio
        if fecha_alta_obj:
            socio.fecha_alta = fecha_alta_obj
        socio.fecha_validez = fecha_validez_obj
        socio.calle = calle
        socio.numero = numero
        socio.piso = piso
        socio.poblacion = poblacion
        
        # Actualizar contraseña si se proporciona
        if password:
            if len(password) < 6:
                flash('La contraseña debe tener al menos 6 caracteres.', 'error')
                from datetime import datetime as dt
                partes_nombre = socio.nombre.split(' ', 2)
                nombre_parts = {
                    'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                    'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                    'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
                }
                return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios, nombre_parts=nombre_parts, datetime=dt)
            socio.set_password(password)
            socio.password_plain = password  # Guardar en texto plano para mostrar a admin
        
        # Procesar beneficiarios
        # Obtener todos los índices de beneficiarios del formulario (pueden no ser secuenciales)
        beneficiarios_indices = set()
        for key in request.form.keys():
            if key.startswith('beneficiario_nombre_'):
                # Extraer el índice del nombre del campo (ej: beneficiario_nombre_1 -> 1)
                try:
                    indice = int(key.replace('beneficiario_nombre_', ''))
                    beneficiarios_indices.add(indice)
                except ValueError:
                    continue
        
        # Validar y preparar nuevos beneficiarios ANTES de eliminar los existentes
        nuevos_beneficiarios = []
        for i in sorted(beneficiarios_indices):
            ben_nombre = request.form.get(f'beneficiario_nombre_{i}', '').strip()
            ben_primer_apellido = request.form.get(f'beneficiario_primer_apellido_{i}', '').strip()
            ben_segundo_apellido = request.form.get(f'beneficiario_segundo_apellido_{i}', '').strip()
            ben_ano = request.form.get(f'beneficiario_ano_{i}', '').strip()
            
            # Validar que los campos obligatorios estén presentes
            if not ben_nombre or not ben_primer_apellido or not ben_ano:
                continue
            
            try:
                ben_ano_nac = int(ben_ano)
                año_actual = datetime.now().year
                if ben_ano_nac < 1900 or ben_ano_nac > año_actual:
                    continue
                
                # Convertir a mayúsculas
                ben_nombre = quitar_acentos(ben_nombre)
                ben_primer_apellido = quitar_acentos(ben_primer_apellido)
                if ben_segundo_apellido:
                    ben_segundo_apellido = quitar_acentos(ben_segundo_apellido)
                
                nuevos_beneficiarios.append({
                    'nombre': ben_nombre,
                    'primer_apellido': ben_primer_apellido,
                    'segundo_apellido': ben_segundo_apellido if ben_segundo_apellido else None,
                    'ano_nacimiento': ben_ano_nac,
                    'indice': i
                })
            except ValueError:
                continue
        
        # Ahora sí, eliminar beneficiarios existentes y crear los nuevos
        for beneficiario in beneficiarios:
            db.session.delete(beneficiario)
        
        # Crear los beneficiarios en la base de datos
        for index, ben_data in enumerate(nuevos_beneficiarios, start=1):
            # Generar número de beneficiario
            numero_beneficiario = f"{socio.numero_socio}-{index}" if socio.numero_socio else None
            
            nuevo_beneficiario = Beneficiario(
                socio_id=socio.id,
                nombre=ben_data['nombre'],
                primer_apellido=ben_data['primer_apellido'],
                segundo_apellido=ben_data['segundo_apellido'],
                ano_nacimiento=ben_data['ano_nacimiento'],
                fecha_validez=socio.fecha_validez,
                numero_beneficiario=numero_beneficiario
            )
            db.session.add(nuevo_beneficiario)
        
        try:
            # Asegurarse de que todos los cambios estén en la sesión
            db.session.add(socio)  # Asegurar que el socio esté en la sesión
            
            # Hacer flush para validar antes del commit
            db.session.flush()
            
            # Commit de todos los cambios
            db.session.commit()
            
            flash(f'Socio {nombre_completo} actualizado exitosamente.', 'success')
            # Redirigir a la página de gestión de socios para evitar problemas de reenvío
            return redirect(url_for('admin.gestion_socios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el socio: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
            import traceback
            traceback.print_exc()
            from datetime import datetime as dt
            # Recargar beneficiarios originales en caso de error
            beneficiarios_originales = Beneficiario.query.filter_by(socio_id=socio.id).order_by(Beneficiario.id).all()
            partes_nombre = socio.nombre.split(' ', 2)
            nombre_parts = {
                'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
                'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
                'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
            }
            return render_template('admin/editar_socio.html', socio=socio, beneficiarios=beneficiarios_originales, nombre_parts=nombre_parts, datetime=dt)
    
    from datetime import datetime as dt
    # Separar nombre completo en partes
    partes_nombre = socio.nombre.split(' ', 2)
    nombre_parts = {
        'nombre': partes_nombre[0] if len(partes_nombre) > 0 else '',
        'primer_apellido': partes_nombre[1] if len(partes_nombre) > 1 else '',
        'segundo_apellido': partes_nombre[2] if len(partes_nombre) > 2 else ''
    }
    
    return render_template('admin/editar_socio.html', 
                         socio=socio, 
                         beneficiarios=beneficiarios, 
                         nombre_parts=nombre_parts,
                         datetime=dt)

@admin_bp.route('/socios/<int:socio_id>/renovar', methods=['GET', 'POST'])
@login_required
@directiva_required
def renovar_socio(socio_id):
    socio = User.query.get_or_404(socio_id)
    
    if request.method == 'POST':
        # Fecha de validez siempre al 31/12 del año en curso
        año_actual = datetime.now().year
        try:
            socio.fecha_validez = datetime(año_actual, 12, 31, 23, 59, 59)
            db.session.commit()
            flash(f'Suscripción de {socio.nombre} renovada exitosamente hasta el 31/12/{año_actual}.', 'success')
            return redirect(url_for('admin.gestion_socios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al renovar la suscripción: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
            import traceback
            traceback.print_exc()
            from datetime import datetime as dt
            return render_template('admin/renovar_socio.html', socio=socio, datetime=dt)
    
    from datetime import datetime as dt
    return render_template('admin/renovar_socio.html', socio=socio, datetime=dt)

@admin_bp.route('/actividades')
@login_required
@directiva_required
def gestion_actividades():
    # Obtener parámetro de búsqueda
    search_query = request.args.get('search', '').strip()
    
    if search_query:
        # Buscar en nombre, descripción o fecha
        actividades = Actividad.query.filter(
            db.or_(
                Actividad.nombre.contains(search_query),
                Actividad.descripcion.contains(search_query),
                db.func.strftime('%d/%m/%Y', Actividad.fecha).contains(search_query)
            )
        ).order_by(Actividad.fecha.desc()).all()
    else:
        actividades = Actividad.query.order_by(Actividad.fecha.desc()).all()
    
    return render_template('admin/actividades.html', actividades=actividades, ahora=datetime.utcnow(), search_query=search_query)

@admin_bp.route('/actividades/nueva', methods=['GET', 'POST'])
@login_required
@directiva_required
def nueva_actividad():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        descripcion = request.form.get('descripcion')
        fecha = request.form.get('fecha')
        aforo_maximo = request.form.get('aforo_maximo')
        edad_minima = request.form.get('edad_minima', '').strip()
        edad_maxima = request.form.get('edad_maxima', '').strip()
        
        # Validaciones
        if not all([nombre, fecha, aforo_maximo]):
            flash('Nombre, fecha y aforo máximo son obligatorios.', 'error')
            return render_template('admin/nueva_actividad.html')
        
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%dT%H:%M')
            aforo = int(aforo_maximo)
            if aforo <= 0:
                raise ValueError()
            
            # Procesar edades (pueden estar vacías)
            edad_min = int(edad_minima) if edad_minima else None
            edad_max = int(edad_maxima) if edad_maxima else None
            
            # Validar que la edad mínima no sea mayor que la máxima
            if edad_min is not None and edad_max is not None and edad_min > edad_max:
                flash('La edad mínima no puede ser mayor que la edad máxima.', 'error')
                return render_template('admin/nueva_actividad.html')
            
            # Validar rangos de edad
            if edad_min is not None and (edad_min < 0 or edad_min > 120):
                flash('La edad mínima debe estar entre 0 y 120 años.', 'error')
                return render_template('admin/nueva_actividad.html')
            
            if edad_max is not None and (edad_max < 0 or edad_max > 120):
                flash('La edad máxima debe estar entre 0 y 120 años.', 'error')
                return render_template('admin/nueva_actividad.html')
                
        except ValueError:
            flash('Fecha, aforo o edades inválidos.', 'error')
            return render_template('admin/nueva_actividad.html')
        
        # Crear nueva actividad
        nueva_actividad = Actividad(
            nombre=nombre,
            descripcion=descripcion,
            fecha=fecha_obj,
            aforo_maximo=aforo,
            edad_minima=edad_min,
            edad_maxima=edad_max
        )
        
        try:
            db.session.add(nueva_actividad)
            db.session.commit()
            flash(f'Actividad "{nombre}" creada exitosamente.', 'success')
            return redirect(url_for('admin.gestion_actividades'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear la actividad: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
            import traceback
            traceback.print_exc()
            return render_template('admin/nueva_actividad.html')
    
    return render_template('admin/nueva_actividad.html')

@admin_bp.route('/actividades/<int:actividad_id>/editar', methods=['GET', 'POST'])
@login_required
@directiva_required
def editar_actividad(actividad_id):
    actividad = Actividad.query.get_or_404(actividad_id)
    
    if request.method == 'POST':
        actividad.nombre = request.form.get('nombre')
        actividad.descripcion = request.form.get('descripcion')
        actividad.aforo_maximo = int(request.form.get('aforo_maximo'))
        edad_minima = request.form.get('edad_minima', '').strip()
        edad_maxima = request.form.get('edad_maxima', '').strip()
        
        try:
            fecha_obj = datetime.strptime(request.form.get('fecha'), '%Y-%m-%dT%H:%M')
            actividad.fecha = fecha_obj
            
            # Procesar edades (pueden estar vacías)
            edad_min = int(edad_minima) if edad_minima else None
            edad_max = int(edad_maxima) if edad_maxima else None
            
            # Validar que la edad mínima no sea mayor que la máxima
            if edad_min is not None and edad_max is not None and edad_min > edad_max:
                flash('La edad mínima no puede ser mayor que la edad máxima.', 'error')
                return render_template('admin/editar_actividad.html', actividad=actividad)
            
            # Validar rangos de edad
            if edad_min is not None and (edad_min < 0 or edad_min > 120):
                flash('La edad mínima debe estar entre 0 y 120 años.', 'error')
                return render_template('admin/editar_actividad.html', actividad=actividad)
            
            if edad_max is not None and (edad_max < 0 or edad_max > 120):
                flash('La edad máxima debe estar entre 0 y 120 años.', 'error')
                return render_template('admin/editar_actividad.html', actividad=actividad)
            
            actividad.edad_minima = edad_min
            actividad.edad_maxima = edad_max
            
        except ValueError:
            flash('Fecha o edades inválidos.', 'error')
            return render_template('admin/editar_actividad.html', actividad=actividad)
        
        try:
            db.session.commit()
            flash(f'Actividad "{actividad.nombre}" actualizada exitosamente.', 'success')
            return redirect(url_for('admin.gestion_actividades'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la actividad: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
            import traceback
            traceback.print_exc()
            return render_template('admin/editar_actividad.html', actividad=actividad)
    
    return render_template('admin/editar_actividad.html', actividad=actividad)

@admin_bp.route('/actividades/<int:actividad_id>/eliminar', methods=['POST'])
@login_required
@directiva_required
def eliminar_actividad(actividad_id):
    actividad = Actividad.query.get_or_404(actividad_id)
    nombre_actividad = actividad.nombre
    
    try:
        db.session.delete(actividad)
        db.session.commit()
        flash(f'Actividad "{nombre_actividad}" eliminada exitosamente.', 'success')
        return redirect(url_for('admin.gestion_actividades'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la actividad: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin.gestion_actividades'))

@admin_bp.route('/actividades/pdf')
@login_required
@directiva_required
def actividades_pdf():
    """Genera un PDF con el listado de todas las actividades"""
    actividades = Actividad.query.order_by(Actividad.fecha.desc()).all()
    ahora = datetime.utcnow()
    
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#333333'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        normal_style = styles['Normal']
        heading_style = styles['Heading2']
        
        story = []
        
        # Título
        story.append(Paragraph("Listado de Actividades", title_style))
        story.append(Paragraph(f"Generado el {ahora.strftime('%d/%m/%Y a las %H:%M')}", 
                              ParagraphStyle('Fecha', parent=normal_style, 
                                           fontSize=10, textColor=colors.grey, 
                                           alignment=TA_CENTER)))
        story.append(Spacer(1, 0.5*cm))
        
        # Información
        story.append(Paragraph(f"<b>Total de actividades:</b> {len(actividades)}", normal_style))
        story.append(Spacer(1, 0.3*cm))
        
        # Tabla de actividades
        if actividades:
            data = [['Actividad', 'Fecha', 'Inscritos', 'Estado']]
            
            for actividad in actividades:
                estado = "Próxima" if actividad.fecha > ahora else "Pasada"
                fecha_str = f"{actividad.fecha.strftime('%d/%m/%Y')}<br/>{actividad.fecha.strftime('%H:%M')}"
                inscritos_str = f"{actividad.numero_inscritos()}/{actividad.aforo_maximo}"
                
                descripcion = actividad.descripcion[:50] + "..." if actividad.descripcion and len(actividad.descripcion) > 50 else (actividad.descripcion or "")
                nombre_completo = f"<b>{actividad.nombre}</b>"
                if descripcion:
                    nombre_completo += f"<br/><i>{descripcion}</i>"
                
                data.append([
                    Paragraph(nombre_completo, normal_style),
                    Paragraph(fecha_str, normal_style),
                    Paragraph(inscritos_str, normal_style),
                    Paragraph(estado, normal_style)
                ])
            
            table = Table(data, colWidths=[7*cm, 3*cm, 3*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No hay actividades registradas.", normal_style))
        
        doc.build(story)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        
    except Exception as e:
        flash(f'No se pudo generar el PDF de actividades: {str(e)}', 'error')
        return redirect(url_for('admin.gestion_actividades'))
    
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=listado_actividades_{datetime.now().strftime("%Y%m%d")}.pdf'
    
    return response

@admin_bp.route('/actividades/<int:actividad_id>/inscritos/pdf')
@login_required
@directiva_required
def inscritos_pdf(actividad_id):
    """Genera un PDF con el listado de inscritos en una actividad"""
    actividad = Actividad.query.get_or_404(actividad_id)
    inscripciones = Inscripcion.query.filter_by(actividad_id=actividad_id).order_by(Inscripcion.fecha_inscripcion).all()
    ahora = datetime.utcnow()
    
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#333333'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        normal_style = styles['Normal']
        heading_style = styles['Heading2']
        
        story = []
        
        # Título
        story.append(Paragraph("Listado de Inscritos", title_style))
        story.append(Paragraph(f"Generado el {ahora.strftime('%d/%m/%Y a las %H:%M')}", 
                              ParagraphStyle('Fecha', parent=normal_style, 
                                           fontSize=10, textColor=colors.grey, 
                                           alignment=TA_CENTER)))
        story.append(Spacer(1, 0.5*cm))
        
        # Información de la actividad
        story.append(Paragraph(f"<b>{actividad.nombre}</b>", heading_style))
        if actividad.descripcion:
            story.append(Paragraph(f"<i>{actividad.descripcion}</i>", normal_style))
        story.append(Paragraph(f"<b>Fecha:</b> {actividad.fecha.strftime('%d/%m/%Y a las %H:%M')}", normal_style))
        story.append(Paragraph(f"<b>Aforo máximo:</b> {actividad.aforo_maximo} personas", normal_style))
        story.append(Spacer(1, 0.3*cm))
        
        # Estadísticas
        asistentes = sum(1 for i in inscripciones if i.asiste)
        no_asistentes = len(inscripciones) - asistentes
        plazas_libres = actividad.plazas_disponibles()
        
        stats_data = [
            ['Total Inscritos', 'Asistieron', 'No Asistieron', 'Plazas Libres'],
            [str(len(inscripciones)), str(asistentes), str(no_asistentes), str(plazas_libres)]
        ]
        stats_table = Table(stats_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Tabla de inscritos
        if inscripciones:
            data = [['#', 'Nombre', 'Nombre de Usuario', 'Fecha de Inscripción', 'Asistencia']]
            
            for idx, inscripcion in enumerate(inscripciones, 1):
                asistencia = "Asistió" if inscripcion.asiste else "No asistió"
                if inscripcion.beneficiario:
                    nombre_completo = f"{inscripcion.beneficiario.nombre} {inscripcion.beneficiario.primer_apellido}"
                    nombre_usuario_mostrar = f"Beneficiario de {inscripcion.usuario.nombre}"
                else:
                    nombre_completo = inscripcion.usuario.nombre
                    nombre_usuario_mostrar = inscripcion.usuario.nombre_usuario
                
                data.append([
                    str(idx),
                    nombre_completo,
                    nombre_usuario_mostrar,
                    inscripcion.fecha_inscripcion.strftime('%d/%m/%Y %H:%M'),
                    asistencia
                ])
            
            table = Table(data, colWidths=[1*cm, 5*cm, 5*cm, 4*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Columna #
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No hay inscripciones para esta actividad.", normal_style))
        
        doc.build(story)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        
    except Exception as e:
        flash(f'No se pudo generar el PDF de inscritos: {str(e)}', 'error')
        return redirect(url_for('admin.ver_inscritos', actividad_id=actividad_id))
    
    response = make_response(pdf_bytes)
    nombre_archivo = f"inscritos_{actividad.nombre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename={nombre_archivo}'
    
    return response

@admin_bp.route('/actividades/<int:actividad_id>/inscritos')
@login_required
@directiva_required
def ver_inscritos(actividad_id):
    actividad = Actividad.query.get_or_404(actividad_id)
    inscripciones = Inscripcion.query.filter_by(actividad_id=actividad_id).all()
    
    from datetime import datetime as dt
    return render_template('admin/inscritos.html', 
                         actividad=actividad, 
                         inscripciones=inscripciones,
                         datetime=dt)

@admin_bp.route('/actividades/<int:actividad_id>/marcar-asistencia/<int:inscripcion_id>', methods=['POST'])
@login_required
@directiva_required
def marcar_asistencia(actividad_id, inscripcion_id):
    inscripcion = Inscripcion.query.get_or_404(inscripcion_id)
    
    # Verificar que la inscripción pertenece a la actividad
    if inscripcion.actividad_id != actividad_id:
        flash('Error: La inscripción no pertenece a esta actividad.', 'error')
        return redirect(url_for('admin.ver_inscritos', actividad_id=actividad_id))
    
    # Cambiar el estado de asistencia
    try:
        inscripcion.asiste = not inscripcion.asiste
        db.session.commit()
        estado = "asistió" if inscripcion.asiste else "no asistió"
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar la asistencia: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin.ver_inscritos', actividad_id=actividad_id))
    if inscripcion.beneficiario:
        nombre_mostrar = f"{inscripcion.beneficiario.nombre} {inscripcion.beneficiario.primer_apellido}"
    else:
        nombre_mostrar = inscripcion.usuario.nombre
    
    flash(f'{nombre_mostrar} marcado como que {estado}.', 'success')
    
    return redirect(url_for('admin.ver_inscritos', actividad_id=actividad_id))

def calcular_nombre_usuario_solicitud(solicitud):
    """Calcula el nombre de usuario predictivo para una solicitud"""
    # Generar nombre de usuario: nombre + iniciales de los dos apellidos + año de nacimiento
    nombre_limpio = solicitud.nombre.lower().replace(' ', '').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
    
    # Obtener iniciales de los apellidos
    inicial_primer_apellido = solicitud.primer_apellido[0].lower() if solicitud.primer_apellido else ''
    inicial_segundo_apellido = solicitud.segundo_apellido[0].lower() if solicitud.segundo_apellido else ''
    
    # Obtener año de nacimiento
    ano_nacimiento = solicitud.fecha_nacimiento.year if solicitud.fecha_nacimiento else ''
    
    nombre_usuario = f"{nombre_limpio}{inicial_primer_apellido}{inicial_segundo_apellido}{ano_nacimiento}"
    
    # Verificar si el nombre de usuario ya existe y generar uno único
    contador = 1
    nombre_usuario_original = nombre_usuario
    while User.query.filter_by(nombre_usuario=nombre_usuario).first():
        nombre_usuario = f"{nombre_limpio}{inicial_primer_apellido}{inicial_segundo_apellido}{ano_nacimiento}{contador}"
        contador += 1
    
    return nombre_usuario

@admin_bp.route('/solicitudes-socios')
@login_required
@directiva_required
def solicitudes_socios():
    """Vista para ver las solicitudes de nuevos socios"""
    estado_filtro = request.args.get('estado', 'por_confirmar')
    
    if estado_filtro == 'todas':
        solicitudes = SolicitudSocio.query.order_by(SolicitudSocio.fecha_solicitud.desc()).all()
    else:
        solicitudes = SolicitudSocio.query.filter_by(estado=estado_filtro).order_by(SolicitudSocio.fecha_solicitud.desc()).all()
    
    # Calcular nombre de usuario para cada solicitud
    solicitudes_con_usuario = []
    for solicitud in solicitudes:
        nombre_usuario = calcular_nombre_usuario_solicitud(solicitud)
        solicitudes_con_usuario.append({
            'solicitud': solicitud,
            'nombre_usuario': nombre_usuario
        })
    
    # Contar por estado
    total_por_confirmar = SolicitudSocio.query.filter_by(estado='por_confirmar').count()
    total_activas = SolicitudSocio.query.filter_by(estado='activa').count()
    total_rechazadas = SolicitudSocio.query.filter_by(estado='rechazada').count()
    
    return render_template('admin/solicitudes_socios.html',
                         solicitudes=solicitudes,
                         solicitudes_con_usuario=solicitudes_con_usuario,
                         estado_filtro=estado_filtro,
                         total_por_confirmar=total_por_confirmar,
                         total_activas=total_activas,
                         total_rechazadas=total_rechazadas)

@admin_bp.route('/solicitudes-confirmadas/excel')
@login_required
@directiva_required
def exportar_solicitudes_confirmadas_excel():
    """Exporta las solicitudes confirmadas a Excel"""
    try:
        # Obtener solo solicitudes confirmadas (estado 'activa')
        solicitudes = SolicitudSocio.query.filter_by(estado='activa').order_by(SolicitudSocio.fecha_confirmacion.desc()).all()
        
        # Crear libro de trabajo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes Confirmadas"
        
        # Estilos para encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'Fecha Solicitud', 'Fecha Confirmación', 'Nombre', 'Primer Apellido', 'Segundo Apellido',
            'Móvil', 'Móvil 2', 'Calle', 'Número', 'Piso', 'Población', 'Dirección Completa',
            'Fecha Nacimiento', 'Miembros Familia', 'Forma de Pago', 'Nombre Usuario', 'Contraseña'
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_num, solicitud in enumerate(solicitudes, 2):
            # Calcular nombre de usuario
            nombre_usuario = calcular_nombre_usuario_solicitud(solicitud)
            
            # Formatear dirección completa
            direccion_completa = ""
            if solicitud.calle and solicitud.numero and solicitud.poblacion:
                direccion_completa = f"{solicitud.calle} {solicitud.numero}"
                if solicitud.piso:
                    direccion_completa += f", {solicitud.piso}"
                direccion_completa += f", {solicitud.poblacion}"
            
            # Formatear fecha de nacimiento
            fecha_nacimiento_str = solicitud.fecha_nacimiento.strftime('%d/%m/%Y') if solicitud.fecha_nacimiento else ''
            
            # Datos de la fila
            row_data = [
                solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_solicitud else '',
                solicitud.fecha_confirmacion.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_confirmacion else '',
                solicitud.nombre,
                solicitud.primer_apellido,
                solicitud.segundo_apellido or '',
                solicitud.movil,
                solicitud.movil2 or '',
                solicitud.calle or '',
                solicitud.numero or '',
                solicitud.piso or '',
                solicitud.poblacion or '',
                direccion_completa,
                fecha_nacimiento_str,
                solicitud.miembros_unidad_familiar,
                solicitud.forma_de_pago,
                nombre_usuario,
                solicitud.password_solicitud or ''
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar ancho de columnas
        column_widths = [18, 18, 15, 15, 15, 12, 12, 20, 8, 10, 15, 40, 15, 12, 12, 20, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'solicitudes_confirmadas_{fecha_str}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error al exportar solicitudes confirmadas a Excel: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin.solicitudes_socios'))

@admin_bp.route('/socios/excel')
@login_required
@directiva_required
def exportar_socios_excel():
    """Exporta todos los socios a Excel"""
    try:
        # Obtener todos los socios
        socios = User.query.filter(User.rol == 'socio').order_by(User.nombre).all()
        
        # Crear libro de trabajo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Socios"
        
        # Estilos para encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'Número Socio', 'Nombre', 'Nombre Usuario', 'Fecha Alta', 'Fecha Validez',
            'Año Nacimiento', 'Fecha Nacimiento', 'Calle', 'Número', 'Piso', 'Población',
            'Dirección Completa', 'Contraseña'
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_num, socio in enumerate(socios, 2):
            # Formatear dirección completa
            direccion_completa = ""
            if socio.calle and socio.numero and socio.poblacion:
                direccion_completa = f"{socio.calle} {socio.numero}"
                if socio.piso:
                    direccion_completa += f", {socio.piso}"
                direccion_completa += f", {socio.poblacion}"
            
            # Formatear fechas
            fecha_alta_str = socio.fecha_alta.strftime('%d/%m/%Y') if socio.fecha_alta else ''
            fecha_validez_str = socio.fecha_validez.strftime('%d/%m/%Y') if socio.fecha_validez else ''
            fecha_nacimiento_str = socio.fecha_nacimiento.strftime('%d/%m/%Y') if socio.fecha_nacimiento else ''
            
            # Datos de la fila
            row_data = [
                socio.numero_socio or '',
                socio.nombre,
                socio.nombre_usuario,
                fecha_alta_str,
                fecha_validez_str,
                socio.ano_nacimiento or '',
                fecha_nacimiento_str,
                socio.calle or '',
                socio.numero or '',
                socio.piso or '',
                socio.poblacion or '',
                direccion_completa,
                socio.password_plain or ''
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar ancho de columnas
        column_widths = [12, 25, 20, 12, 12, 12, 12, 20, 8, 10, 15, 40, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'socios_{fecha_str}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error al exportar socios a Excel: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin.gestion_socios'))

@admin_bp.route('/solicitudes-socios/<int:solicitud_id>')
@login_required
@directiva_required
def ver_solicitud(solicitud_id):
    """Vista para ver el detalle de una solicitud"""
    solicitud = SolicitudSocio.query.get_or_404(solicitud_id)
    beneficiarios = BeneficiarioSolicitud.query.filter_by(solicitud_id=solicitud_id).order_by(BeneficiarioSolicitud.id).all()
    
    return render_template('admin/ver_solicitud.html',
                         solicitud=solicitud,
                         beneficiarios=beneficiarios)

@admin_bp.route('/solicitudes-socios/<int:solicitud_id>/editar', methods=['GET', 'POST'])
@login_required
@directiva_required
def editar_solicitud(solicitud_id):
    """Vista para editar una solicitud"""
    solicitud = SolicitudSocio.query.get_or_404(solicitud_id)
    
    # Solo se puede editar si está por confirmar
    if solicitud.estado != 'por_confirmar':
        flash('Solo se pueden editar solicitudes pendientes de confirmación.', 'error')
        return redirect(url_for('admin.ver_solicitud', solicitud_id=solicitud_id))
    
    beneficiarios = BeneficiarioSolicitud.query.filter_by(solicitud_id=solicitud_id).order_by(BeneficiarioSolicitud.id).all()
    
    if request.method == 'POST':
        from datetime import datetime as dt
        
        try:
            # Obtener datos del formulario
            nuevo_nombre = request.form.get('nombre', '').strip().upper()
            nuevo_primer_apellido = request.form.get('primer_apellido', '').strip().upper()
            nuevo_segundo_apellido = request.form.get('segundo_apellido', '').strip().upper() or None
            nuevo_movil = request.form.get('movil', '').strip()
            nuevo_movil2 = request.form.get('movil2', '').strip()  # Segundo móvil (opcional)
            nuevos_miembros = int(request.form.get('miembros_unidad_familiar', 1))
            nueva_forma_pago = request.form.get('forma_de_pago', '').strip()
            
            # Validar móvil
            if not re.match(r'^\d{9}$', nuevo_movil):
                flash('El número de móvil debe tener 9 dígitos.', 'error')
                beneficiarios = BeneficiarioSolicitud.query.filter_by(solicitud_id=solicitud_id).order_by(BeneficiarioSolicitud.id).all()
                return render_template('admin/editar_solicitud.html', solicitud=solicitud, beneficiarios=beneficiarios, datetime=dt)
            
            # Validar móvil2 si está presente
            if nuevo_movil2 and not re.match(r'^\d{9}$', nuevo_movil2):
                flash('El segundo número de móvil debe tener 9 dígitos o estar vacío.', 'error')
                beneficiarios = BeneficiarioSolicitud.query.filter_by(solicitud_id=solicitud_id).order_by(BeneficiarioSolicitud.id).all()
                return render_template('admin/editar_solicitud.html', solicitud=solicitud, beneficiarios=beneficiarios, datetime=dt)
            
            # Normalizar movil2 (None si está vacío)
            nuevo_movil2 = nuevo_movil2 if nuevo_movil2 else None
            
            # Actualizar datos del socio directamente en el objeto
            solicitud.nombre = nuevo_nombre
            solicitud.primer_apellido = nuevo_primer_apellido
            solicitud.segundo_apellido = nuevo_segundo_apellido
            solicitud.movil = nuevo_movil
            solicitud.movil2 = nuevo_movil2
            solicitud.miembros_unidad_familiar = nuevos_miembros
            solicitud.forma_de_pago = nueva_forma_pago
            
            # Asegurar que SQLAlchemy detecte los cambios
            db.session.add(solicitud)
            
            # Actualizar o crear beneficiarios
            nuevos_beneficiarios_count = nuevos_miembros - 1
            
            # Eliminar beneficiarios existentes
            if beneficiarios:
                for beneficiario in beneficiarios:
                    db.session.delete(beneficiario)
            
            # Crear nuevos beneficiarios
            if nuevos_beneficiarios_count > 0:
                for i in range(1, nuevos_beneficiarios_count + 1):
                    beneficiario_nombre = request.form.get(f'beneficiario_nombre_{i}', '').strip().upper()
                    beneficiario_primer_apellido = request.form.get(f'beneficiario_primer_apellido_{i}', '').strip().upper()
                    beneficiario_segundo_apellido = request.form.get(f'beneficiario_segundo_apellido_{i}', '').strip().upper() or None
                    beneficiario_ano = request.form.get(f'beneficiario_ano_{i}', '').strip()
                    
                    if beneficiario_nombre and beneficiario_primer_apellido and beneficiario_ano:
                        try:
                            ano_nacimiento = int(beneficiario_ano)
                            año_actual = datetime.now().year
                            if ano_nacimiento < 1900 or ano_nacimiento > año_actual:
                                flash(f'El año de nacimiento del beneficiario {i} no es válido.', 'error')
                                db.session.rollback()
                                beneficiarios = BeneficiarioSolicitud.query.filter_by(solicitud_id=solicitud_id).order_by(BeneficiarioSolicitud.id).all()
                                return render_template('admin/editar_solicitud.html', solicitud=solicitud, beneficiarios=beneficiarios, datetime=dt)
                            
                            nuevo_beneficiario = BeneficiarioSolicitud(
                                solicitud_id=solicitud.id,
                                nombre=beneficiario_nombre,
                                primer_apellido=beneficiario_primer_apellido,
                                segundo_apellido=beneficiario_segundo_apellido,
                                ano_nacimiento=ano_nacimiento
                            )
                            db.session.add(nuevo_beneficiario)
                        except ValueError:
                            flash(f'El año de nacimiento del beneficiario {i} debe ser un número válido.', 'error')
                            db.session.rollback()
                            beneficiarios = BeneficiarioSolicitud.query.filter_by(solicitud_id=solicitud_id).order_by(BeneficiarioSolicitud.id).all()
                            return render_template('admin/editar_solicitud.html', solicitud=solicitud, beneficiarios=beneficiarios, datetime=dt)
            
            # Commit de TODOS los cambios en una sola transacción
            db.session.commit()
            
            # Cerrar la sesión actual y crear una nueva para asegurar que se recarguen los datos
            db.session.close()
            
            # Recargar la solicitud desde la BD para verificar que se guardó
            solicitud_verificada = SolicitudSocio.query.get(solicitud_id)
            if not solicitud_verificada:
                raise Exception("No se pudo verificar que la solicitud se guardó correctamente")
            
            # Verificar que los valores se guardaron correctamente
            if solicitud_verificada.nombre != nuevo_nombre:
                raise Exception(f"Los cambios no se guardaron. Esperado: {nuevo_nombre}, Obtenido: {solicitud_verificada.nombre}")
            
            flash('Solicitud actualizada correctamente.', 'success')
            return redirect(url_for('admin.ver_solicitud', solicitud_id=solicitud_id))
            
        except Exception as e:
            db.session.rollback()
            error_msg = str(e)
            flash(f'Error al actualizar la solicitud: {error_msg}. Por favor, inténtalo de nuevo.', 'error')
            import traceback
            print(f"[ERROR] Error al actualizar solicitud {solicitud_id}:")
            traceback.print_exc()
            # Recargar la solicitud original desde la base de datos
            db.session.expire_all()
            solicitud = SolicitudSocio.query.get_or_404(solicitud_id)
            beneficiarios = BeneficiarioSolicitud.query.filter_by(solicitud_id=solicitud_id).order_by(BeneficiarioSolicitud.id).all()
            return render_template('admin/editar_solicitud.html', solicitud=solicitud, beneficiarios=beneficiarios, datetime=dt)
    
    from datetime import datetime as dt
    return render_template('admin/editar_solicitud.html', solicitud=solicitud, beneficiarios=beneficiarios, datetime=dt)

@admin_bp.route('/solicitudes-socios/<int:solicitud_id>/confirmar', methods=['POST'])
@login_required
@directiva_required
def confirmar_solicitud(solicitud_id):
    """Confirmar una solicitud y crear el usuario"""
    solicitud = SolicitudSocio.query.get_or_404(solicitud_id)
    
    if solicitud.estado != 'por_confirmar':
        flash('Esta solicitud ya ha sido procesada.', 'error')
        return redirect(url_for('admin.solicitudes_socios'))
    
    # Generar número de socio (0001, 0002, etc.)
    ultimo_socio = User.query.filter(User.numero_socio.isnot(None)).order_by(User.numero_socio.desc()).first()
    if ultimo_socio and ultimo_socio.numero_socio:
        try:
            ultimo_numero = int(ultimo_socio.numero_socio)
            nuevo_numero = ultimo_numero + 1
        except ValueError:
            nuevo_numero = 1
    else:
        nuevo_numero = 1
    
    numero_socio = f"{nuevo_numero:04d}"  # Formato 0001, 0002, etc.
    
    # Generar nombre completo
    nombre_completo = f"{solicitud.nombre} {solicitud.primer_apellido}"
    if solicitud.segundo_apellido:
        nombre_completo += f" {solicitud.segundo_apellido}"
    
    # Generar nombre de usuario: nombre + iniciales de los dos apellidos + año de nacimiento
    # Limpiar y normalizar nombre
    nombre_limpio = solicitud.nombre.lower().replace(' ', '').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
    
    # Obtener iniciales de los apellidos
    inicial_primer_apellido = solicitud.primer_apellido[0].lower() if solicitud.primer_apellido else ''
    inicial_segundo_apellido = solicitud.segundo_apellido[0].lower() if solicitud.segundo_apellido else ''
    
    # Obtener año de nacimiento
    ano_nacimiento = solicitud.fecha_nacimiento.year if solicitud.fecha_nacimiento else ''
    
    nombre_usuario = f"{nombre_limpio}{inicial_primer_apellido}{inicial_segundo_apellido}{ano_nacimiento}"
    
    # Verificar si el nombre de usuario ya existe y generar uno único
    contador = 1
    nombre_usuario_original = nombre_usuario
    while User.query.filter_by(nombre_usuario=nombre_usuario).first():
        nombre_usuario = f"{nombre_limpio}{inicial_primer_apellido}{inicial_segundo_apellido}{ano_nacimiento}{contador}"
        contador += 1
    
    # Usar la contraseña de la solicitud
    password = solicitud.password_solicitud
    if not password:
        # Si no hay contraseña en la solicitud, generar una temporal
        password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
    
    # Crear usuario
    # Fecha de validez siempre al 31/12 del año en curso
    año_actual = datetime.now().year
    fecha_validez = datetime(año_actual, 12, 31, 23, 59, 59)
    
    # Calcular año de nacimiento desde fecha_nacimiento
    ano_nacimiento = None
    if solicitud.fecha_nacimiento:
        ano_nacimiento = solicitud.fecha_nacimiento.year
    
    nuevo_socio = User(
        nombre=nombre_completo,
        nombre_usuario=nombre_usuario,
        rol='socio',
        fecha_alta=datetime.utcnow(),
        fecha_validez=fecha_validez,
        numero_socio=numero_socio,
        fecha_nacimiento=solicitud.fecha_nacimiento,
        ano_nacimiento=ano_nacimiento,
        password_plain=password,  # Guardar contraseña en texto plano para mostrar a admin
        calle=solicitud.calle,
        numero=solicitud.numero,
        piso=solicitud.piso,
        poblacion=solicitud.poblacion
    )
    nuevo_socio.set_password(password)
    
    # Actualizar solicitud
    solicitud.estado = 'activa'
    solicitud.fecha_confirmacion = datetime.utcnow()
    
    try:
        db.session.add(nuevo_socio)
        db.session.flush()  # Para obtener el ID del nuevo socio
        
        # Crear beneficiarios asociados al socio con números
        beneficiarios_solicitud = BeneficiarioSolicitud.query.filter_by(solicitud_id=solicitud.id).all()
        for index, beneficiario_solicitud in enumerate(beneficiarios_solicitud, start=1):
            numero_beneficiario = f"{numero_socio}-{index}"  # Formato 0001-1, 0001-2, etc.
            beneficiario = Beneficiario(
                socio_id=nuevo_socio.id,
                nombre=beneficiario_solicitud.nombre,
                primer_apellido=beneficiario_solicitud.primer_apellido,
                segundo_apellido=beneficiario_solicitud.segundo_apellido,
                ano_nacimiento=beneficiario_solicitud.ano_nacimiento,
                fecha_validez=fecha_validez,  # Misma fecha de vigencia que el socio
                numero_beneficiario=numero_beneficiario
            )
            db.session.add(beneficiario)
        
        db.session.commit()
        beneficiarios_count = len(beneficiarios_solicitud)
        mensaje = f'Solicitud confirmada. Usuario creado: {nombre_usuario} (Número de socio: {numero_socio}) con contraseña: {password}'
        if beneficiarios_count > 0:
            mensaje += f'. Se crearon {beneficiarios_count} beneficiario(s).'
        
        flash(mensaje, 'success')
        return redirect(url_for('admin.solicitudes_socios'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al confirmar la solicitud: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
        return redirect(url_for('admin.ver_solicitud', solicitud_id=solicitud_id))

@admin_bp.route('/solicitudes-socios/<int:solicitud_id>/rechazar', methods=['POST'])
@login_required
@directiva_required
def rechazar_solicitud(solicitud_id):
    """Rechazar una solicitud"""
    solicitud = SolicitudSocio.query.get_or_404(solicitud_id)
    
    if solicitud.estado != 'por_confirmar':
        flash('Esta solicitud ya ha sido procesada.', 'error')
        return redirect(url_for('admin.solicitudes_socios'))
    
    solicitud.estado = 'rechazada'
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error al rechazar la solicitud: {str(e)}. Por favor, inténtalo de nuevo.', 'error')
        return redirect(url_for('admin.ver_solicitud', solicitud_id=solicitud_id))
    
    flash('Solicitud rechazada.', 'info')
    return redirect(url_for('admin.solicitudes_socios'))

@admin_bp.route('/exportar-datos', methods=['GET'])
@login_required
@directiva_required
def exportar_datos():
    """Exporta todos los datos de la base de datos a un archivo JSON"""
    try:
        # Recopilar todos los datos
        datos = {
            'fecha_exportacion': datetime.utcnow().isoformat(),
            'version': '1.0',
            'usuarios': [],
            'actividades': [],
            'inscripciones': [],
            'beneficiarios': [],
            'solicitudes_socio': [],
            'beneficiarios_solicitud': []
        }
        
        # Exportar usuarios
        usuarios = User.query.all()
        for usuario in usuarios:
            datos['usuarios'].append({
                'id': usuario.id,
                'nombre': usuario.nombre,
                'nombre_usuario': usuario.nombre_usuario,
                'password_hash': usuario.password_hash,
                'password_plain': usuario.password_plain,
                'rol': usuario.rol,
                'fecha_alta': usuario.fecha_alta.isoformat() if usuario.fecha_alta else None,
                'fecha_validez': usuario.fecha_validez.isoformat() if usuario.fecha_validez else None,
                'ano_nacimiento': usuario.ano_nacimiento,
                'fecha_nacimiento': usuario.fecha_nacimiento.isoformat() if usuario.fecha_nacimiento else None,
                'numero_socio': usuario.numero_socio,
                'calle': usuario.calle,
                'numero': usuario.numero,
                'piso': usuario.piso,
                'poblacion': usuario.poblacion
            })
        
        # Exportar actividades
        actividades = Actividad.query.all()
        for actividad in actividades:
            datos['actividades'].append({
                'id': actividad.id,
                'nombre': actividad.nombre,
                'descripcion': actividad.descripcion,
                'fecha': actividad.fecha.isoformat() if actividad.fecha else None,
                'aforo_maximo': actividad.aforo_maximo,
                'edad_minima': actividad.edad_minima,
                'edad_maxima': actividad.edad_maxima,
                'fecha_creacion': actividad.fecha_creacion.isoformat() if actividad.fecha_creacion else None
            })
        
        # Exportar inscripciones
        inscripciones = Inscripcion.query.all()
        for inscripcion in inscripciones:
            datos['inscripciones'].append({
                'id': inscripcion.id,
                'user_id': inscripcion.user_id,
                'actividad_id': inscripcion.actividad_id,
                'beneficiario_id': inscripcion.beneficiario_id,
                'fecha_inscripcion': inscripcion.fecha_inscripcion.isoformat() if inscripcion.fecha_inscripcion else None,
                'asiste': inscripcion.asiste
            })
        
        # Exportar beneficiarios
        beneficiarios = Beneficiario.query.all()
        for beneficiario in beneficiarios:
            datos['beneficiarios'].append({
                'id': beneficiario.id,
                'socio_id': beneficiario.socio_id,
                'nombre': beneficiario.nombre,
                'primer_apellido': beneficiario.primer_apellido,
                'segundo_apellido': beneficiario.segundo_apellido,
                'ano_nacimiento': beneficiario.ano_nacimiento,
                'fecha_validez': beneficiario.fecha_validez.isoformat() if beneficiario.fecha_validez else None,
                'numero_beneficiario': beneficiario.numero_beneficiario
            })
        
        # Exportar solicitudes
        solicitudes = SolicitudSocio.query.all()
        for solicitud in solicitudes:
            datos['solicitudes_socio'].append({
                'id': solicitud.id,
                'nombre': solicitud.nombre,
                'primer_apellido': solicitud.primer_apellido,
                'segundo_apellido': solicitud.segundo_apellido,
                'movil': solicitud.movil,
                'fecha_nacimiento': solicitud.fecha_nacimiento.isoformat() if solicitud.fecha_nacimiento else None,
                'miembros_unidad_familiar': solicitud.miembros_unidad_familiar,
                'forma_de_pago': solicitud.forma_de_pago,
                'estado': solicitud.estado,
                'fecha_solicitud': solicitud.fecha_solicitud.isoformat() if solicitud.fecha_solicitud else None,
                'fecha_confirmacion': solicitud.fecha_confirmacion.isoformat() if solicitud.fecha_confirmacion else None,
                'password_solicitud': solicitud.password_solicitud,
                'calle': solicitud.calle,
                'numero': solicitud.numero,
                'piso': solicitud.piso,
                'poblacion': solicitud.poblacion
            })
        
        # Exportar beneficiarios de solicitudes
        beneficiarios_solicitud = BeneficiarioSolicitud.query.all()
        for ben_sol in beneficiarios_solicitud:
            datos['beneficiarios_solicitud'].append({
                'id': ben_sol.id,
                'solicitud_id': ben_sol.solicitud_id,
                'nombre': ben_sol.nombre,
                'primer_apellido': ben_sol.primer_apellido,
                'segundo_apellido': ben_sol.segundo_apellido,
                'ano_nacimiento': ben_sol.ano_nacimiento
            })
        
        # Convertir a JSON
        json_data = json.dumps(datos, indent=2, ensure_ascii=False)
        
        # Crear archivo en memoria
        output = BytesIO()
        output.write(json_data.encode('utf-8'))
        output.seek(0)
        
        # Generar nombre de archivo con fecha
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'backup_asociacion_{fecha_str}.txt'
        
        return send_file(
            output,
            mimetype='text/plain',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error al exportar los datos: {str(e)}', 'error')
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/importar-datos', methods=['GET', 'POST'])
@login_required
@directiva_required
def importar_datos():
    """Importa datos desde un archivo JSON"""
    if request.method == 'GET':
        return render_template('admin/importar_datos.html')
    
    if 'archivo' not in request.files:
        flash('No se ha seleccionado ningún archivo.', 'error')
        return render_template('admin/importar_datos.html')
    
    archivo = request.files['archivo']
    if archivo.filename == '':
        flash('No se ha seleccionado ningún archivo.', 'error')
        return render_template('admin/importar_datos.html')
    
    try:
        # Leer el archivo
        contenido = archivo.read().decode('utf-8')
        datos = json.loads(contenido)
        
        # Validar estructura
        if 'version' not in datos:
            flash('El archivo no tiene el formato correcto.', 'error')
            return render_template('admin/importar_datos.html')
        
        # Preguntar si se debe limpiar la base de datos primero
        limpiar_bd = request.form.get('limpiar_bd') == 'on'
        
        if limpiar_bd:
            # Eliminar todos los datos existentes (en orden inverso de dependencias)
            BeneficiarioSolicitud.query.delete()
            Beneficiario.query.delete()
            Inscripcion.query.delete()
            SolicitudSocio.query.delete()
            Actividad.query.delete()
            User.query.delete()
            db.session.commit()
        
        # Importar usuarios
        usuarios_importados = 0
        for user_data in datos.get('usuarios', []):
            try:
                # Verificar si el usuario ya existe (compatibilidad con datos antiguos que usan 'email')
                nombre_usuario = user_data.get('nombre_usuario') or user_data.get('email')
                if not nombre_usuario:
                    flash('Usuario sin nombre_usuario, saltando.', 'warning')
                    continue
                    
                if not limpiar_bd and User.query.filter_by(nombre_usuario=nombre_usuario).first():
                    flash(f"Usuario con nombre_usuario {nombre_usuario} ya existe, saltando.", 'warning')
                    continue
                
                usuario = User(
                    nombre=user_data['nombre'],
                    nombre_usuario=nombre_usuario,
                    password_hash=user_data['password_hash'],
                    password_plain=user_data.get('password_plain'),
                    rol=user_data['rol'],
                    fecha_alta=datetime.fromisoformat(user_data['fecha_alta']) if user_data.get('fecha_alta') else datetime.utcnow(),
                    fecha_validez=datetime.fromisoformat(user_data['fecha_validez']) if user_data.get('fecha_validez') else datetime.utcnow(),
                    ano_nacimiento=user_data.get('ano_nacimiento'),
                    fecha_nacimiento=datetime.fromisoformat(user_data['fecha_nacimiento']).date() if user_data.get('fecha_nacimiento') else None,
                    numero_socio=user_data.get('numero_socio'),
                    calle=user_data.get('calle'),
                    numero=user_data.get('numero'),
                    piso=user_data.get('piso'),
                    poblacion=user_data.get('poblacion')
                )
                db.session.add(usuario)
                usuarios_importados += 1
            except Exception as e:
                flash(f'Error al importar usuario {user_data.get("nombre_usuario", user_data.get("email", "desconocido"))}: {str(e)}', 'warning')
                continue
        
        # Importar actividades
        actividades_importadas = 0
        for act_data in datos.get('actividades', []):
            try:
                actividad = Actividad(
                    nombre=act_data['nombre'],
                    descripcion=act_data.get('descripcion'),
                    fecha=datetime.fromisoformat(act_data['fecha']) if act_data.get('fecha') else datetime.utcnow(),
                    aforo_maximo=act_data['aforo_maximo'],
                    edad_minima=act_data.get('edad_minima'),
                    edad_maxima=act_data.get('edad_maxima'),
                    fecha_creacion=datetime.fromisoformat(act_data['fecha_creacion']) if act_data.get('fecha_creacion') else datetime.utcnow()
                )
                db.session.add(actividad)
                actividades_importadas += 1
            except Exception as e:
                flash(f'Error al importar actividad {act_data.get("nombre", "desconocida")}: {str(e)}', 'warning')
                continue
        
        # Importar beneficiarios (después de usuarios)
        beneficiarios_importados = 0
        for ben_data in datos.get('beneficiarios', []):
            try:
                # Verificar que el socio exista
                if not User.query.get(ben_data['socio_id']):
                    continue
                
                beneficiario = Beneficiario(
                    socio_id=ben_data['socio_id'],
                    nombre=ben_data['nombre'],
                    primer_apellido=ben_data['primer_apellido'],
                    segundo_apellido=ben_data.get('segundo_apellido'),
                    ano_nacimiento=ben_data['ano_nacimiento'],
                    fecha_validez=datetime.fromisoformat(ben_data['fecha_validez']) if ben_data.get('fecha_validez') else datetime.utcnow(),
                    numero_beneficiario=ben_data.get('numero_beneficiario')
                )
                db.session.add(beneficiario)
                beneficiarios_importados += 1
            except Exception as e:
                flash(f'Error al importar beneficiario: {str(e)}', 'warning')
                continue
        
        # Importar inscripciones (después de usuarios y actividades)
        inscripciones_importadas = 0
        for ins_data in datos.get('inscripciones', []):
            try:
                # Verificar que el usuario y la actividad existan
                if not User.query.get(ins_data['user_id']):
                    continue
                if not Actividad.query.get(ins_data['actividad_id']):
                    continue
                if ins_data.get('beneficiario_id') and not Beneficiario.query.get(ins_data['beneficiario_id']):
                    continue
                
                inscripcion = Inscripcion(
                    user_id=ins_data['user_id'],
                    actividad_id=ins_data['actividad_id'],
                    beneficiario_id=ins_data.get('beneficiario_id'),
                    fecha_inscripcion=datetime.fromisoformat(ins_data['fecha_inscripcion']) if ins_data.get('fecha_inscripcion') else datetime.utcnow(),
                    asiste=ins_data.get('asiste', False)
                )
                db.session.add(inscripcion)
                inscripciones_importadas += 1
            except Exception as e:
                flash(f'Error al importar inscripción: {str(e)}', 'warning')
                continue
        
        # Importar solicitudes
        solicitudes_importadas = 0
        for sol_data in datos.get('solicitudes_socio', []):
            try:
                solicitud = SolicitudSocio(
                    nombre=sol_data['nombre'],
                    primer_apellido=sol_data['primer_apellido'],
                    segundo_apellido=sol_data.get('segundo_apellido'),
                    movil=sol_data['movil'],
                    fecha_nacimiento=datetime.fromisoformat(sol_data['fecha_nacimiento']).date() if sol_data.get('fecha_nacimiento') else None,
                    miembros_unidad_familiar=sol_data['miembros_unidad_familiar'],
                    forma_de_pago=sol_data['forma_de_pago'],
                    estado=sol_data['estado'],
                    fecha_solicitud=datetime.fromisoformat(sol_data['fecha_solicitud']) if sol_data.get('fecha_solicitud') else datetime.utcnow(),
                    fecha_confirmacion=datetime.fromisoformat(sol_data['fecha_confirmacion']) if sol_data.get('fecha_confirmacion') else None,
                    password_solicitud=sol_data.get('password_solicitud'),
                    calle=sol_data.get('calle'),
                    numero=sol_data.get('numero'),
                    piso=sol_data.get('piso'),
                    poblacion=sol_data.get('poblacion')
                )
                db.session.add(solicitud)
                db.session.flush()  # Para obtener el ID
                
                # Importar beneficiarios de solicitudes
                for ben_sol_data in datos.get('beneficiarios_solicitud', []):
                    if ben_sol_data.get('solicitud_id') == sol_data.get('id'):
                        ben_sol = BeneficiarioSolicitud(
                            solicitud_id=solicitud.id,
                            nombre=ben_sol_data['nombre'],
                            primer_apellido=ben_sol_data['primer_apellido'],
                            segundo_apellido=ben_sol_data.get('segundo_apellido'),
                            ano_nacimiento=ben_sol_data['ano_nacimiento']
                        )
                        db.session.add(ben_sol)
                
                solicitudes_importadas += 1
            except Exception as e:
                flash(f'Error al importar solicitud: {str(e)}', 'warning')
                continue
        
        # Commit final con manejo de errores
        try:
            db.session.commit()
            flash(f'Importación completada: {usuarios_importados} usuarios, {actividades_importadas} actividades, {beneficiarios_importados} beneficiarios, {inscripciones_importadas} inscripciones, {solicitudes_importadas} solicitudes.', 'success')
            return redirect(url_for('admin.dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar los datos importados: {str(e)}. Todos los cambios han sido revertidos.', 'error')
            import traceback
            traceback.print_exc()
            return render_template('admin/importar_datos.html')
        
    except json.JSONDecodeError:
        flash('El archivo no es un JSON válido.', 'error')
        return render_template('admin/importar_datos.html')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al importar los datos: {str(e)}', 'error')
        return render_template('admin/importar_datos.html')

@admin_bp.route('/descargar-base-datos', methods=['GET'])
@login_required
@directiva_required
def descargar_base_datos():
    """Descarga la base de datos completa (SQLite como .db, PostgreSQL como dump SQL)"""
    try:
        database_url = current_app.config.get('SQLALCHEMY_DATABASE_URI', 'sqlite:///asociacion.db')
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # SQLite - descargar archivo .db directamente
        if 'sqlite' in database_url.lower():
            db_path = database_url.replace('sqlite:///', '')
            if not os.path.isabs(db_path):
                # Ruta relativa, buscar en instance/ primero, luego en el directorio raíz
                instance_path = os.path.join(current_app.instance_path, db_path)
                root_path = db_path
                
                if os.path.exists(instance_path):
                    db_path = instance_path
                elif os.path.exists(root_path):
                    db_path = root_path
                else:
                    # Si no existe en ninguna ubicación, intentar con instance/
                    db_path = instance_path
            
            if not os.path.exists(db_path):
                flash(f'No se encontró el archivo de base de datos SQLite en: {db_path}', 'error')
                return redirect(url_for('admin.dashboard'))
            
            # Cerrar todas las conexiones y hacer checkpoint de WAL para asegurar consistencia
            try:
                db.session.close_all()
                db.engine.dispose()
                
                # Si está en modo WAL, hacer checkpoint para asegurar que todos los cambios están en el archivo principal
                from sqlalchemy import text
                with db.engine.connect() as conn:
                    conn.execute(text('PRAGMA wal_checkpoint(FULL);'))
                    conn.commit()
                
                # Cerrar de nuevo después del checkpoint
                db.session.close_all()
                db.engine.dispose()
            except Exception as e:
                print(f"[WARNING] No se pudo hacer checkpoint de WAL: {e}")
            
            # Leer el archivo completo y enviarlo
            with open(db_path, 'rb') as f:
                db_data = f.read()
            
            filename = f'backup_bd_completa_{fecha_str}.db'
            output = BytesIO(db_data)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/x-sqlite3',
                as_attachment=True,
                download_name=filename
            )
        
        # PostgreSQL - generar dump SQL
        elif 'postgres' in database_url.lower():
            # Extraer información de la URL
            # Formato: postgresql://user:password@host:port/database
            try:
                from urllib.parse import urlparse
                parsed = urlparse(database_url)
                
                db_name = parsed.path.lstrip('/')
                db_user = parsed.username
                db_password = parsed.password
                db_host = parsed.hostname
                db_port = parsed.port or 5432
                
                # Generar dump usando pg_dump
                filename = f'backup_bd_completa_{fecha_str}.sql'
                
                # Comando pg_dump
                env = os.environ.copy()
                env['PGPASSWORD'] = db_password
                
                cmd = [
                    'pg_dump',
                    '-h', db_host,
                    '-p', str(db_port),
                    '-U', db_user,
                    '-d', db_name,
                    '--no-owner',
                    '--no-acl',
                    '--clean',
                    '--if-exists'
                ]
                
                # Verificar si pg_dump está disponible
                try:
                    subprocess.run(['pg_dump', '--version'], capture_output=True, timeout=5)
                except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
                    flash('pg_dump no está disponible en el sistema. La descarga de PostgreSQL requiere que pg_dump esté instalado.', 'error')
                    return redirect(url_for('admin.dashboard'))
                
                # Ejecutar pg_dump
                try:
                    result = subprocess.run(
                        cmd,
                        env=env,
                        capture_output=True,
                        text=True,
                        timeout=300  # 5 minutos máximo
                    )
                except subprocess.TimeoutExpired:
                    flash('La operación de dump tardó demasiado tiempo. Inténtalo de nuevo.', 'error')
                    return redirect(url_for('admin.dashboard'))
                except FileNotFoundError:
                    flash('pg_dump no está disponible en el sistema. La descarga de PostgreSQL requiere que pg_dump esté instalado.', 'error')
                    return redirect(url_for('admin.dashboard'))
                
                if result.returncode != 0:
                    error_msg = result.stderr if result.stderr else 'Error desconocido al generar dump'
                    flash(f'Error al generar dump de PostgreSQL: {error_msg}', 'error')
                    return redirect(url_for('admin.dashboard'))
                
                # Crear archivo en memoria
                output = BytesIO()
                output.write(result.stdout.encode('utf-8'))
                output.seek(0)
                
                return send_file(
                    output,
                    mimetype='application/sql',
                    as_attachment=True,
                    download_name=filename
                )
                
            except Exception as e:
                flash(f'Error al procesar la base de datos PostgreSQL: {str(e)}', 'error')
                return redirect(url_for('admin.dashboard'))
        
        else:
            flash('Tipo de base de datos no soportado para descarga completa.', 'error')
            return redirect(url_for('admin.dashboard'))
            
    except Exception as e:
        flash(f'Error al descargar la base de datos: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/importar-base-datos', methods=['POST'])
@login_required
@directiva_required
def importar_base_datos():
    """Importa/restaura la base de datos desde un archivo SQLite subido"""
    # Verificar que solo jmurillo puede usar esta función
    if current_user.nombre_usuario != 'jmurillo':
        flash('No tienes permisos para realizar esta operación.', 'error')
        return redirect(url_for('admin.dashboard'))
    
    if 'archivo' not in request.files:
        flash('No se ha seleccionado ningún archivo.', 'error')
        return redirect(url_for('admin.dashboard'))
    
    archivo = request.files['archivo']
    if archivo.filename == '':
        flash('No se ha seleccionado ningún archivo.', 'error')
        return redirect(url_for('admin.dashboard'))
    
    try:
        database_url = current_app.config.get('SQLALCHEMY_DATABASE_URI', 'sqlite:///asociacion.db')
        
        # SQLite - importar desde archivo .db
        if 'sqlite' in database_url.lower():
            db_path = database_url.replace('sqlite:///', '')
            if not os.path.isabs(db_path):
                # Buscar en instance/ primero, luego en el directorio raíz
                instance_path = os.path.join(current_app.instance_path, db_path)
                root_path = db_path
                
                if os.path.exists(instance_path):
                    db_path = instance_path
                elif os.path.exists(root_path):
                    db_path = root_path
                else:
                    # Si no existe, usar instance/ como ubicación por defecto
                    db_path = instance_path
            
            # Cerrar todas las conexiones antes de importar
            db.session.close_all()
            db.engine.dispose()
            
            # Leer el archivo subido
            archivo_data = archivo.read()
            
            # Validar que el archivo no esté vacío
            if len(archivo_data) == 0:
                flash('El archivo de base de datos está vacío.', 'error')
                return redirect(url_for('admin.dashboard'))
            
            # Validar que sea un archivo SQLite válido (debe empezar con "SQLite format 3")
            if not archivo_data.startswith(b'SQLite format 3\x00'):
                flash('El archivo no parece ser un archivo SQLite válido.', 'error')
                return redirect(url_for('admin.dashboard'))
            
            # Hacer backup del archivo actual antes de importar
            if os.path.exists(db_path):
                backup_actual = f"{db_path}.backup_antes_importacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                try:
                    shutil.copy2(db_path, backup_actual)
                    flash(f'Se creó un backup del archivo actual en: {backup_actual}', 'info')
                except Exception as e:
                    flash(f'Advertencia: No se pudo crear backup del archivo actual: {e}', 'warning')
            
            # Eliminar archivos auxiliares de WAL si existen (para evitar inconsistencias)
            archivos_auxiliares = [
                f"{db_path}-shm",
                f"{db_path}-wal",
                f"{db_path}.shm",
                f"{db_path}.wal"
            ]
            for archivo_aux in archivos_auxiliares:
                if os.path.exists(archivo_aux):
                    try:
                        os.remove(archivo_aux)
                    except Exception as e:
                        print(f"[WARNING] No se pudo eliminar archivo auxiliar {archivo_aux}: {e}")
            
            # Asegurar que el directorio existe
            db_dir = os.path.dirname(db_path)
            if db_dir and not os.path.exists(db_dir):
                os.makedirs(db_dir, exist_ok=True)
            
            # Escribir el nuevo archivo completo
            with open(db_path, 'wb') as f:
                f.write(archivo_data)
            
            # Verificar que el archivo se escribió correctamente
            if not os.path.exists(db_path) or os.path.getsize(db_path) != len(archivo_data):
                flash('Error al escribir el archivo de base de datos. La importación puede estar incompleta.', 'error')
                return redirect(url_for('admin.dashboard'))
            
            # Reiniciar conexiones de SQLAlchemy
            db.session.close_all()
            db.engine.dispose()
            
            flash('Base de datos SQLite importada exitosamente. Por favor, recarga la página para ver los cambios.', 'success')
            return redirect(url_for('admin.dashboard'))
        
        else:
            flash('Tipo de base de datos no soportado para importación completa.', 'error')
            return redirect(url_for('admin.dashboard'))
            
    except Exception as e:
        flash(f'Error al importar la base de datos: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/restaurar-base-datos', methods=['GET', 'POST'])
@login_required
@directiva_required
def restaurar_base_datos():
    """Restaura la base de datos desde un archivo de backup"""
    if request.method == 'GET':
        return render_template('admin/restaurar_base_datos.html')
    
    # Verificar confirmación
    confirmacion = request.form.get('confirmacion', '').strip()
    if confirmacion != 'RESTAURAR':
        flash('Debes escribir "RESTAURAR" en mayúsculas para confirmar la operación.', 'error')
        return render_template('admin/restaurar_base_datos.html')
    
    if 'archivo' not in request.files:
        flash('No se ha seleccionado ningún archivo.', 'error')
        return render_template('admin/restaurar_base_datos.html')
    
    archivo = request.files['archivo']
    if archivo.filename == '':
        flash('No se ha seleccionado ningún archivo.', 'error')
        return render_template('admin/restaurar_base_datos.html')
    
    try:
        database_url = current_app.config.get('SQLALCHEMY_DATABASE_URI', 'sqlite:///asociacion.db')
        
        # SQLite - restaurar desde archivo .db
        if 'sqlite' in database_url.lower():
            db_path = database_url.replace('sqlite:///', '')
            if not os.path.isabs(db_path):
                db_path = os.path.join(current_app.instance_path, db_path)
            
            # Cerrar todas las conexiones antes de restaurar
            db.session.close_all()
            db.engine.dispose()
            
            # Leer el archivo subido
            archivo_data = archivo.read()
            
            # Validar que el archivo no esté vacío
            if len(archivo_data) == 0:
                flash('El archivo de backup está vacío.', 'error')
                return render_template('admin/restaurar_base_datos.html')
            
            # Validar que sea un archivo SQLite válido (debe empezar con "SQLite format 3")
            if not archivo_data.startswith(b'SQLite format 3\x00'):
                flash('El archivo no parece ser un archivo SQLite válido.', 'error')
                return render_template('admin/restaurar_base_datos.html')
            
            # Hacer backup del archivo actual antes de restaurar
            if os.path.exists(db_path):
                backup_actual = f"{db_path}.backup_antes_restauracion_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                try:
                    shutil.copy2(db_path, backup_actual)
                    flash(f'Se creó un backup del archivo actual en: {backup_actual}', 'info')
                except Exception as e:
                    flash(f'Advertencia: No se pudo crear backup del archivo actual: {e}', 'warning')
            
            # Eliminar archivos auxiliares de WAL si existen (para evitar inconsistencias)
            archivos_auxiliares = [
                f"{db_path}-shm",
                f"{db_path}-wal",
                f"{db_path}.shm",
                f"{db_path}.wal"
            ]
            for archivo_aux in archivos_auxiliares:
                if os.path.exists(archivo_aux):
                    try:
                        os.remove(archivo_aux)
                    except Exception as e:
                        print(f"[WARNING] No se pudo eliminar archivo auxiliar {archivo_aux}: {e}")
            
            # Asegurar que el directorio existe
            db_dir = os.path.dirname(db_path)
            if db_dir and not os.path.exists(db_dir):
                os.makedirs(db_dir, exist_ok=True)
            
            # Escribir el nuevo archivo completo
            with open(db_path, 'wb') as f:
                f.write(archivo_data)
            
            # Verificar que el archivo se escribió correctamente
            if not os.path.exists(db_path) or os.path.getsize(db_path) != len(archivo_data):
                flash('Error al escribir el archivo de base de datos. La restauración puede estar incompleta.', 'error')
                return render_template('admin/restaurar_base_datos.html')
            
            # Reiniciar conexiones de SQLAlchemy
            db.session.close_all()
            db.engine.dispose()
            
            flash('Base de datos SQLite restaurada exitosamente. La aplicación se reiniciará.', 'success')
            return redirect(url_for('admin.dashboard'))
        
        # PostgreSQL - restaurar desde dump SQL
        elif 'postgres' in database_url.lower():
            try:
                from urllib.parse import urlparse
                parsed = urlparse(database_url)
                
                db_name = parsed.path.lstrip('/')
                db_user = parsed.username
                db_password = parsed.password
                db_host = parsed.hostname
                db_port = parsed.port or 5432
                
                # Leer el contenido del archivo SQL
                contenido_sql = archivo.read().decode('utf-8')
                
                # Verificar si psql está disponible
                try:
                    subprocess.run(['psql', '--version'], capture_output=True, timeout=5)
                except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
                    flash('psql no está disponible en el sistema. La restauración de PostgreSQL requiere que psql esté instalado.', 'error')
                    return render_template('admin/restaurar_base_datos.html')
                
                # Ejecutar el dump usando psql
                env = os.environ.copy()
                env['PGPASSWORD'] = db_password
                
                cmd = [
                    'psql',
                    '-h', db_host,
                    '-p', str(db_port),
                    '-U', db_user,
                    '-d', db_name,
                    '-f', '-'  # Leer desde stdin
                ]
                
                try:
                    result = subprocess.run(
                        cmd,
                        input=contenido_sql,
                        env=env,
                        capture_output=True,
                        text=True,
                        timeout=600  # 10 minutos máximo
                    )
                except subprocess.TimeoutExpired:
                    flash('La operación de restauración tardó demasiado tiempo. Inténtalo de nuevo.', 'error')
                    return render_template('admin/restaurar_base_datos.html')
                except FileNotFoundError:
                    flash('psql no está disponible en el sistema. La restauración de PostgreSQL requiere que psql esté instalado.', 'error')
                    return render_template('admin/restaurar_base_datos.html')
                
                if result.returncode != 0:
                    error_msg = result.stderr if result.stderr else 'Error desconocido al restaurar'
                    flash(f'Error al restaurar PostgreSQL: {error_msg}', 'error')
                    return render_template('admin/restaurar_base_datos.html')
                
                # Reiniciar conexiones de SQLAlchemy
                db.session.close_all()
                db.engine.dispose()
                
                flash('Base de datos PostgreSQL restaurada exitosamente.', 'success')
                return redirect(url_for('admin.dashboard'))
                
            except Exception as e:
                flash(f'Error al restaurar PostgreSQL: {str(e)}', 'error')
                import traceback
                traceback.print_exc()
                return render_template('admin/restaurar_base_datos.html')
        
        else:
            flash('Tipo de base de datos no soportado para restauración.', 'error')
            return render_template('admin/restaurar_base_datos.html')
            
    except Exception as e:
        flash(f'Error al restaurar la base de datos: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return render_template('admin/restaurar_base_datos.html')
